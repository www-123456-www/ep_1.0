# !/usr/bin/env python
# -*- coding:utf-8 -*-
# @Author  : wyy
# @file       :  to_openpyxl.py
# @Time    : 2021/11/12 16:02
# @Function:
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def pandas_to_openpyxl(df, sheet_name):
    """"将df  转成 openpyxl
    :param df:      df
    :param sheet_name:
    :return:     sheet
    """
    sheet = Workbook().create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=False):
        sheet.append(r)
    return sheet


def copy_exlce(old_path, new_path):
    """

    :param old_path:
    :param new_path:
    :param df_s:  [df]  某天的 df (多个sheet,多个df)
    """
    table = load_workbook(old_path)
    table.save(new_path)
    table.close()
    # print('copy_.')


def fill_data(old_path, new_path, df, sheet_name):  #############################???????????
    copy_exlce(old_path, new_path)


def openpyxl_cell(obj, x, y, sheet):
    """
    :param obj:  pandas_to_openpyxl(df , sheetname)  = sheet
    :param x:  初始行位置
    :param y:  初始列位置
    sheet  :load_workbook(new_path)[sheetname]
    :return:
    """
    data_l = obj.iter_rows()  # 将文件转成 迭代器
    for a_, a in enumerate(data_l):
        for b_, b in enumerate(a):
            # print("(a_+x ,, b_+y )",(a_+y+1,b_+x+1))
            sheet.cell(row=(a_ + x), column=(b_ + y), value=b.value)  # obj的数据填充到目标文件


def write_day(x, y, day, sheet):
    sheet.cell(row=x, column=y, value=day)  # obj的数据填充到目标文件


if __name__ == '__main__':
    table = Workbook()
    sheet = table.active
    sheet.cell(row=6, column=2, value="www")
    print("www")
    table.save("wyy.xlsx")
