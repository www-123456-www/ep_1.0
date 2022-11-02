# 这是一个示例 Python 脚本。

# 按 Shift+F10 执行或将其替换为您的代码。
# 按 双击 Shift 在所有地方搜索类、文件、工具窗口、操作和设置。
import datetime
import logging
import time
import pandas as pp
from all.dao.read_config import get_StationEncode_s, get_days, get_params, get_Sheetnames, get_df_If, \
    get_StationEncode_time, get_name
from all.dao.to_openpyxl import pandas_to_openpyxl, copy_exlce, openpyxl_cell, write_day
from all.dao.to_pandas import data_to_daysData, set_index_if
from all.log import logger
from all.server.login import get_data
from openpyxl import load_workbook
import os

time_data=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
user=get_name()[0]
pwd=get_name()[1]

def convert_to_day(day):
    """
    :param day: str(day)  "%Y%m%d"
    :return:
    """
    day_str=str(day)
    timeArray = time.strptime(day_str,"%Y%m%d")
    otherStyleTime = time.strftime("%Y/%m/%d", timeArray)
    return otherStyleTime

if __name__ == '__main__':
    # file = open('d:/ab.txt', 'w')
    #side=get_StationEncode_s()
    side = get_StationEncode_time()

    str_log="需要生成巡回表的电站："+str(side)
    logger.info(str_log)
    # create_str_to_txt(str_log)

    for StationEncode in side:
        str_log="正生成 %s 的文件" % StationEncode
        logger.info(str_log)
        # create_str_to_txt(str_log)
        old_path = "config/%s.xlsx" % (StationEncode)
        day = get_days(StationEncode)
        from_list = get_data(user, pwd, StationEncode, day[0], day[1])
        str_log="时间段"+ str(from_list[1])
        logger.info(str_log)
        for table in from_list[1]:  #时间
            logger.info("时间啊",table)
            new_path="file/%s_%s.xlsx" %(StationEncode,str(table))
            copy_exlce(old_path, new_path)
            file=load_workbook(new_path)
            df = data_to_daysData(from_list[0], table)#
            df[1].astype(dtype="object")
            logger.info(str(StationEncode+" : "+str(table)))
            logger.info(str(df))
            for sheetname in get_Sheetnames(StationEncode):
                logger.info("sheet",sheetname)
                params = get_params(StationEncode, sheetname)
                df_if=set_index_if(df,get_df_If(params))
                df_new = pp.DataFrame(df_if).T.loc[1:26]  # 筛选 最后的结果每个sheet
                logger.info(sheetname)
                logger.info(str(df_new))
                logger.info(df_new)
                new_data=pandas_to_openpyxl(df_new, sheetname)
                sheet=file[sheetname]
                openpyxl_cell(new_data,6,2,sheet)
                table_new=convert_to_day(table)
                sheet.cell(row=1,column=1, value=table_new)
            file.save(new_path)
            file.close()

            str_log=str(new_path)+" 文件生成成功"
            logger.info(str_log)
            # create_str_to_txt(str_log)
    time.sleep(5)
# 访问 https://www.jetbrains.com/help/pycharm/ 获取 PyCharm 帮助
